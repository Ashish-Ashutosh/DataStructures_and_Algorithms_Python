from _datetime import datetime
import openpyxl
import sqlite3
with sqlite3.connect('Expenses.sqlite') as connection:
   cursor = connection.cursor()

#class Expenses seems to work
class Expense:
    def __init__(self, category, amount, date):
        self.category = category
        self.amount = amount
        self.date = date
        cursor.execute("INSERT INTO Expense (Amount, CategoryId,Date) VALUES (?,?,?);", (amount,category,date))
        rows = cursor.fetchall()
        connection.commit()
        for row in rows:
            print(row)
        return

def export_to_excel():
    #today: str = datetime.today().strftime('%Y-%m-%d')
    #print(today)
    date = datetime.now().date()
    workbook = openpyxl.Workbook(f"{date}.xlsx")
    workbook.create_sheet("Expenses", 0)
    worksheet = workbook.active
    worksheet.title = "Expenses"
    worksheet.append(['Amount', 'Category', 'Date'])
    workbook.save(f"{date}.xlsx")

    #"Category" -  Column2 of Excel sheet (open excel sheet , write it and then close the excel sheet)
    workbook = openpyxl.load_workbook(filename=f"{date}.xlsx")
    worksheet = workbook.active
    categories = getCategory()
    counter = 2
    for category in categories:
        to_string = ''.join(category)
        worksheet.cell(row=counter, column=2).value = to_string
        counter = counter + 1
    workbook.save(f"{date}.xlsx")

    #"Amount" - Column1 of Excel Sheet (open excel sheet , write it and then close the excel sheet)
    workbook = openpyxl.load_workbook(filename=f"{date}.xlsx")
    worksheet = workbook.active
    amounts = getAmount()
    counter = 2
    for amount in amounts:
        number = ''.join(map(str, amount))
        worksheet.cell(row=counter, column=1).value = number
        counter = counter + 1
    workbook.save(f"{date}.xlsx")

    #"Date" - Column3 of Excel sheet (open excel sheet , write it and then close the excel sheet)
    workbook = openpyxl.load_workbook(filename=f"{date}.xlsx")
    worksheet = workbook.active
    dates_of_purchase = getDate()
    counter = 2
    for dates in dates_of_purchase:
        to_string_date = ''.join(dates)
        worksheet.cell(row=counter, column=3).value = to_string_date
        counter = counter + 1
    workbook.save(f"{date}.xlsx")
    main()


def main():
    answer = int(input('What do you want to do?\n 1 - add expense\n 2 - export expenses to Excel\n 3 - Exit\n >>'))
    if answer == 1:
        cursor.execute('SELECT * FROM Category')
        rows = cursor.fetchall()
        for row in rows:
            print(row)
        question_category = int(input('To which category belongs this expense? Choose number > '))
        category = question_category
        question_amount = input('What is the amount of the expense? > ')
        amount = question_amount
        date = datetime.now().date()
        entry = Expense(category, amount, date)
        main()
#till here everything seems to work (task 1 with sqlite)
    if answer == 2:
        entry = export_to_excel()
    else:
        pass


def getCategory():
    cursor.execute("SELECT Category.Name FROM Category INNER JOIN Expense on Expense.CategoryId = Category.CategoryID")
    rows = cursor.fetchall()
    return rows

def getAmount():
    cursor.execute("SELECT Expense.Amount FROM Expense")
    rows = cursor.fetchall()
    print(len(rows))
    return rows

def getDate():
    cursor.execute("SELECT Expense.Date FROM Expense")
    rows = cursor.fetchall()
    print(len(rows))
    return rows



main()

